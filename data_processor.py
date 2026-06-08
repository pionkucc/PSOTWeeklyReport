"""
数据处理模块
负责读取Excel数据并进行预处理
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from config import INPUT_FILE, PANEL_SHEET_INDEX, OVERDUE_DAYS, REWORK_THRESHOLD


def load_data():
    """读取Excel数据，返回 DataFrame 和统计数据"""
    df = pd.read_excel(INPUT_FILE)
    total = len(df)

    # 从sheet2读取任务项统计
    df_sheet2 = pd.read_excel(INPUT_FILE, sheet_name=1)
    task_count = len(df_sheet2) - 1
    subtask_count = df_sheet2.iloc[-1, 4] if len(df_sheet2) > 0 else 0

    return df, total, task_count, subtask_count


def preprocess_data(df):
    """预处理数据，修复周期列类型转换"""
    df['缺陷修复周期(天)'] = pd.to_numeric(
        df['缺陷修复周期(天)'].replace('/', np.nan), errors='coerce'
    )
    return df


def load_panels_data():
    """读取sheet3公共面板数据，保留富文本格式"""
    wb = load_workbook(INPUT_FILE)
    ws = wb.worksheets[PANEL_SHEET_INDEX]

    panels = []
    for row_idx in range(2, ws.max_row + 1):
        title = ws.cell(row=row_idx, column=1).value
        content_cell = ws.cell(row=row_idx, column=2)

        if not title:  # 跳过空标题行
            continue

        # 解析内容，保留富文本格式
        content_parts = []
        content_value = content_cell.value

        if content_value:
            # 检查是否是富文本（CellRichText类型）
            if hasattr(content_value, '__iter__') and not isinstance(content_value, str):
                for part in content_value:
                    part_text = ''
                    part_bold = False
                    part_color = None

                    if isinstance(part, str):
                        part_text = part
                    elif hasattr(part, 'text'):
                        part_text = part.text or ''
                        if hasattr(part, 'font') and part.font:
                            font = part.font
                            part_bold = getattr(font, 'b', False) or False
                            if hasattr(font, 'color') and font.color:
                                if hasattr(font.color, 'rgb') and font.color.rgb:
                                    part_color = str(font.color.rgb)

                    content_parts.append({
                        'text': part_text,
                        'bold': part_bold,
                        'color': part_color
                    })
            else:
                # 纯文本，检查单元格整体字体样式
                cell_bold = False
                cell_color = None
                if content_cell.font:
                    cell_bold = getattr(content_cell.font, 'b', False) or False
                    if hasattr(content_cell.font, 'color') and content_cell.font.color:
                        if hasattr(content_cell.font.color, 'rgb') and content_cell.font.color.rgb:
                            cell_color = str(content_cell.font.color.rgb)

                content_parts.append({
                    'text': str(content_value),
                    'bold': cell_bold,
                    'color': cell_color
                })

        # 读取所有图片路径（从第3列开始的所有列）
        import os
        base_dir = os.path.dirname(os.path.abspath(INPUT_FILE))
        images = []  # 图片列表，按列顺序存储

        for col_idx in range(3, ws.max_column + 1):
            image_cell = ws.cell(row=row_idx, column=col_idx)
            if image_cell.value:
                cell_val = str(image_cell.value).strip()
                # 跳过Excel公式（如 =DISPIMG(...)）
                if not cell_val.startswith('='):
                    # 处理路径
                    # 注意：Excel中以/开头的路径应视为相对于项目根目录的路径，而非绝对路径
                    if cell_val.startswith('/'):
                        cell_val = cell_val[1:]
                    full_path = os.path.join(base_dir, cell_val)
                    if os.path.isfile(full_path):
                        images.append(full_path)

        panels.append({
            'title': str(title),
            'content_parts': content_parts,
            'images': images  # 图片列表
        })

    wb.close()
    return panels


def load_sheet2_data():
    """读取Sheet2测试进度和缺陷统计数据"""
    try:
        df = pd.read_excel(INPUT_FILE, sheet_name=1)
        return df
    except Exception as e:
        print(f'读取Sheet2数据失败: {e}')
        return None


def load_warning_data():
    """
    读取缺陷预警数据

    返回:
        dict: {
            'overdue_rework': 超期返工数据列表,
            'overdue': 超期数据列表,
            'rework': 返工数据列表,
            'total': 总预警数
        }
        每条数据包含:
            - type: 类型标识
            - code: 缺陷编号
            - handler: 处理人员
            - summary: 缺陷摘要
            - overdue_days: 超期天数
            - rework_count: 返工次数
            - full_data: 全部字段字典（用于弹窗）
    """
    from datetime import datetime

    try:
        df = pd.read_excel(INPUT_FILE, sheet_name=0)
        now = datetime.now()

        overdue_rework_list = []
        overdue_list = []
        rework_list = []
        rework_codes = set()  # 记录已加入超期返工的缺陷编号，避免重复

        # 筛选超期返工数据（优先级最高）
        # 条件：状态=ReOpen + 返工次数>=阈值 + 有修复时间 + (当前时间-修复时间)>=阈值
        overdue_rework_mask = (
            (df['缺陷状态'] == 'ReOpen') &
            (df['返工次数'] >= REWORK_THRESHOLD) &
            (df['修复时间'].notna())
        )
        for _, row in df[overdue_rework_mask].iterrows():
            fix_time = pd.to_datetime(row['修复时间'])
            days_overdue = (now - fix_time).days
            if days_overdue >= OVERDUE_DAYS:
                rework_codes.add(row['缺陷编号'])
                overdue_rework_list.append({
                    'type': 'overdue_rework',
                    'code': row['缺陷编号'],
                    'handler': row['处理人员'] if pd.notna(row['处理人员']) else '',
                    'summary': row['缺陷摘要'] if pd.notna(row['缺陷摘要']) else '',
                    'overdue_days': days_overdue,
                    'rework_count': int(row['返工次数']),
                    'full_data': row.to_dict()
                })

        # 按超期天数倒序
        overdue_rework_list.sort(key=lambda x: (x['overdue_days'], x['rework_count']), reverse=True)

        # 筛选超期数据
        # 条件：状态=New + (当前时间-登记时间)>=阈值
        overdue_mask = (df['缺陷状态'] == 'New') & (df['登记时间.1'].notna())
        for _, row in df[overdue_mask].iterrows():
            reg_time = pd.to_datetime(row['登记时间.1'])
            days_overdue = (now - reg_time).days
            if days_overdue >= OVERDUE_DAYS:
                overdue_list.append({
                    'type': 'overdue',
                    'code': row['缺陷编号'],
                    'handler': row['处理人员'] if pd.notna(row['处理人员']) else '',
                    'summary': row['缺陷摘要'] if pd.notna(row['缺陷摘要']) else '',
                    'overdue_days': days_overdue,
                    'rework_count': 0,
                    'full_data': row.to_dict()
                })

        # 按超期天数倒序
        overdue_list.sort(key=lambda x: x['overdue_days'], reverse=True)

        # 筛选返工数据（排除已在超期返工中的）
        # 条件：状态=ReOpen + 返工次数>=阈值 + 未在超期返工列表中
        rework_mask = (
            (df['缺陷状态'] == 'ReOpen') &
            (df['返工次数'] >= REWORK_THRESHOLD) &
            (~df['缺陷编号'].isin(rework_codes))
        )
        for _, row in df[rework_mask].iterrows():
            rework_list.append({
                'type': 'rework',
                'code': row['缺陷编号'],
                'handler': row['处理人员'] if pd.notna(row['处理人员']) else '',
                'summary': row['缺陷摘要'] if pd.notna(row['缺陷摘要']) else '',
                'overdue_days': 0,
                'rework_count': int(row['返工次数']),
                'full_data': row.to_dict()
            })

        # 按返工次数倒序
        rework_list.sort(key=lambda x: x['rework_count'], reverse=True)

        total = len(overdue_rework_list) + len(overdue_list) + len(rework_list)

        return {
            'overdue_rework': overdue_rework_list,
            'overdue': overdue_list,
            'rework': rework_list,
            'total': total
        }

    except Exception as e:
        print(f'读取预警数据失败: {e}')
        return {'overdue_rework': [], 'overdue': [], 'rework': [], 'total': 0}
